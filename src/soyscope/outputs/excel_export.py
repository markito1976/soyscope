"""Generate Excel workbooks from the SoyScope database."""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from ..db import Database

logger = logging.getLogger(__name__)

# ── Shared style constants ──────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_MAX_COL_WIDTH = 60


class ExcelExporter:
    """Exports SoyScope data to a styled .xlsx workbook."""

    def __init__(self, db: Database, output_dir: Path) -> None:
        self.db = db
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ── public entry point ──────────────────────────────────────────

    def export(self, filename: str | None = None) -> Path:
        """Create a full report workbook and return the saved file path."""
        if filename is None:
            date_str = datetime.now().strftime("%Y%m%d")
            filename = f"soyscope_report_{date_str}.xlsx"

        wb = Workbook()

        # Sheet 1 - Master List (the default sheet created by Workbook())
        self._build_master_list(wb.active)

        # Sheet 2 - Sector Matrix
        self._build_sector_matrix(wb.create_sheet("Sector Matrix"))

        # Sheet 3 - Timeline
        self._build_timeline(wb.create_sheet("Timeline"))

        # Sheet 4 - Top Novel
        self._build_top_novel(wb.create_sheet("Top Novel"))

        # Sheet 5 - Commercial
        self._build_commercial(wb.create_sheet("Commercial"))

        # Sheet 6 - Statistics
        self._build_statistics(wb.create_sheet("Statistics"))

        dest = self.output_dir / filename
        wb.save(str(dest))
        logger.info("Excel report saved to %s", dest)
        return dest

    # ── Sheet builders ──────────────────────────────────────────────

    def _build_master_list(self, ws) -> None:
        """Sheet 1: every finding, one row per record."""
        ws.title = "Master List"
        headers = [
            "ID", "Title", "Year", "DOI", "Source API",
            "Source Type", "Venue", "Citation Count", "OA Status", "URL",
        ]
        ws.append(headers)

        findings = self.db.get_all_findings()
        for f in findings:
            ws.append([
                f.get("id"),
                f.get("title"),
                f.get("year"),
                f.get("doi"),
                f.get("source_api"),
                f.get("source_type"),
                f.get("venue"),
                f.get("citation_count"),
                f.get("open_access_status"),
                f.get("url"),
            ])

        self._style_header(ws)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # Force Title column to 60; let the rest auto-fit
        self._auto_width(ws)
        ws.column_dimensions["B"].width = 60

    def _build_sector_matrix(self, ws) -> None:
        """Sheet 2: pivot of sectors x derivatives with finding counts."""
        stats = self.db.get_stats()
        matrix_rows = stats.get("sector_derivative_matrix", [])

        # Collect unique sectors and derivatives preserving order of appearance
        sectors_seen: dict[str, None] = {}
        derivatives_seen: dict[str, None] = {}
        lookup: dict[tuple[str, str], int] = {}
        for row in matrix_rows:
            sec = row["sector"]
            der = row["derivative"]
            cnt = row["count"]
            sectors_seen.setdefault(sec, None)
            derivatives_seen.setdefault(der, None)
            lookup[(sec, der)] = cnt

        sectors = list(sectors_seen)
        derivatives = list(derivatives_seen)

        # Header row: blank corner + derivative names
        ws.append(["Sector \\ Derivative"] + derivatives)

        # Data rows
        for sec in sectors:
            row_data: list[str | int] = [sec]
            for der in derivatives:
                row_data.append(lookup.get((sec, der), 0))
            ws.append(row_data)

        # Colour-code data cells (gradient white -> green)
        max_count = max(lookup.values()) if lookup else 1
        for row_idx in range(2, len(sectors) + 2):
            for col_idx in range(2, len(derivatives) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value if cell.value else 0
                intensity = int(255 - (val / max_count) * 200) if max_count else 255
                green_hex = f"00{intensity:02X}00"
                # RGB fill: red channel = intensity, green = 255, blue = intensity
                fill_color = f"{intensity:02X}FF{intensity:02X}"
                cell.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid",
                )
                cell.alignment = Alignment(horizontal="center")

        self._style_header(ws)
        self._auto_width(ws)

    def _build_timeline(self, ws) -> None:
        """Sheet 3: findings count by year with a bar chart."""
        stats = self.db.get_stats()
        by_year = stats.get("by_year", {})

        ws.append(["Year", "Count"])
        for year in sorted(by_year):
            ws.append([year, by_year[year]])

        self._style_header(ws)
        self._auto_width(ws)

        # Bar chart
        if by_year:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Findings by Year"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Year"
            chart.style = 10

            num_rows = len(by_year)
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=num_rows + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=num_rows + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            chart.width = 20
            chart.height = 12

            ws.add_chart(chart, "D2")

    def _build_top_novel(self, ws) -> None:
        """Sheet 4: top 100 findings ranked by novelty score."""
        headers = [
            "Rank", "Title", "Year", "Novelty Score",
            "TRL", "Status", "Summary",
        ]
        ws.append(headers)

        with self.db.connect() as conn:
            rows = conn.execute(
                """SELECT e.novelty_score, e.trl_estimate,
                          e.commercialization_status, e.ai_summary,
                          f.title, f.year
                   FROM enrichments e
                   JOIN findings f ON e.finding_id = f.id
                   WHERE e.novelty_score IS NOT NULL
                   ORDER BY e.novelty_score DESC
                   LIMIT 100""",
            ).fetchall()

        for rank, r in enumerate(rows, start=1):
            ws.append([
                rank,
                r["title"],
                r["year"],
                r["novelty_score"],
                r["trl_estimate"],
                r["commercialization_status"],
                r["ai_summary"],
            ])

        self._style_header(ws)
        self._auto_width(ws)

    def _build_commercial(self, ws) -> None:
        """Sheet 5: findings with commercial / scaling / mature status."""
        headers = [
            "Title", "Year", "Status", "TRL",
            "Key Players", "Summary",
        ]
        ws.append(headers)

        with self.db.connect() as conn:
            rows = conn.execute(
                """SELECT f.title, f.year,
                          e.commercialization_status, e.trl_estimate,
                          e.key_players, e.ai_summary
                   FROM enrichments e
                   JOIN findings f ON e.finding_id = f.id
                   WHERE e.commercialization_status IN ('commercial', 'scaling', 'mature')
                   ORDER BY e.trl_estimate DESC""",
            ).fetchall()

        for r in rows:
            key_players_raw = r["key_players"]
            if key_players_raw:
                try:
                    players = json.loads(key_players_raw)
                    if isinstance(players, list):
                        key_players_str = ", ".join(str(p) for p in players)
                    else:
                        key_players_str = str(players)
                except (json.JSONDecodeError, TypeError):
                    key_players_str = str(key_players_raw)
            else:
                key_players_str = ""

            ws.append([
                r["title"],
                r["year"],
                r["commercialization_status"],
                r["trl_estimate"],
                key_players_str,
                r["ai_summary"],
            ])

        self._style_header(ws)
        self._auto_width(ws)

    def _build_statistics(self, ws) -> None:
        """Sheet 6: database-level statistics as key-value pairs."""
        stats = self.db.get_stats()

        ws.append(["Metric", "Value"])

        ws.append(["Total Findings", stats.get("total_findings", 0)])
        ws.append(["Total Sectors", stats.get("total_sectors", 0)])
        ws.append(["Total Derivatives", stats.get("total_derivatives", 0)])
        ws.append(["Total Enriched", stats.get("total_enriched", 0)])
        ws.append(["Total Checkoff Projects", stats.get("total_checkoff", 0)])
        ws.append(["Total Tags", stats.get("total_tags", 0)])
        ws.append(["Total Search Runs", stats.get("total_runs", 0)])

        # Enrichment coverage breakdown
        ws.append(["", ""])
        ws.append(["Enrichment Coverage", ""])
        ws.append(["  Catalog Tier", stats.get("enrichment_catalog", 0)])
        ws.append(["  Summary Tier", stats.get("enrichment_summary", 0)])
        ws.append(["  Deep Tier", stats.get("enrichment_deep", 0)])

        # By source
        ws.append(["", ""])
        ws.append(["Findings by Source API", ""])
        for source, count in stats.get("by_source", {}).items():
            ws.append([f"  {source}", count])

        # By type
        ws.append(["", ""])
        ws.append(["Findings by Source Type", ""])
        for stype, count in stats.get("by_type", {}).items():
            ws.append([f"  {stype}", count])

        self._style_header(ws)
        self._auto_width(ws)

    # ── Helper methods ──────────────────────────────────────────────

    def _style_header(self, ws, row: int = 1) -> None:
        """Apply blue-fill, white-bold styling to the header row."""
        for cell in ws[row]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

    def _auto_width(self, ws) -> None:
        """Set each column width to fit the longest cell value, capped at *_MAX_COL_WIDTH*."""
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except (TypeError, AttributeError):
                    pass
            # Add a small padding
            adjusted = min(max_length + 3, _MAX_COL_WIDTH)
            ws.column_dimensions[col_letter].width = adjusted
